import re
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

try:
    from scipy.interpolate import CubicSpline
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False


def find_latest_csv(data_dir: Path) -> Path:
    """Возвращает самый свежий IVtrace_*.csv файл в data_dir."""
    data_dir = Path(data_dir)
    if not data_dir.exists():
        raise FileNotFoundError(f"Папка {data_dir} не найдена. Сначала выполните измерения.")

    csv_files = list(data_dir.glob("IVtrace_*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"Нет CSV-файлов IVtrace_*.csv в папке {data_dir}.")

    return max(csv_files, key=lambda f: f.stat().st_mtime)


def _read_metadata(csv_path: Path) -> dict:
    """Читает строки, начинающиеся с '#', в виде 'ключ: значение'."""
    metadata = {}
    with open(csv_path, 'r', encoding='utf-8') as f:
        for line in f:
            if not line.startswith('#'):
                break
            match = re.match(r'#\s*(.*?)\s*:\s*(.*)', line)
            if match:
                key, value = match.groups()
                metadata[key.strip()] = value.strip()
    return metadata


def load_and_analyze(latest_csv: Path, I_nom: float, X: float, save_png: bool = True,
                     show: bool = False, close_fig: bool = True,
                     show_error_labels: bool = False,
                     xlim: Optional[tuple] = None, y1lim: Optional[tuple] = None,
                     y2lim: Optional[tuple] = None) -> dict:
    """
    Читает CSV с результатами измерения, считает приведённую погрешность
    относительно ожидаемого выходного тока датчика (коэффициент 1:X),
    строит график (амплитудная характеристика + погрешность) и сохраняет PNG
    рядом с CSV.

    show_error_labels — подписывать каждую точку нижнего графика её
    погрешностью (п.30), вертикально (чтобы не налезали друг на друга при
    частом шаге), с точностью 0.001 %.
    xlim/y1lim/y2lim — необязательные (min, max) для оси X (общая для обоих
    графиков) и осей Y верхнего/нижнего графика (п.36); None — авто (matplotlib
    сам подберёт диапазон, как раньше).

    Возвращает словарь со статистикой и путями к файлам.
    """
    latest_csv = Path(latest_csv)
    metadata = _read_metadata(latest_csv)
    df = pd.read_csv(latest_csv, comment='#')

    if df.empty:
        raise ValueError(f"Файл {latest_csv} не содержит данных измерений.")

    label = metadata.get('Датчик', 'Неизвестный датчик')
    # 'Branch' присутствует в новых файлах (forward/reverse через реле).
    # Для старых файлов без реле (одна ветвь, метаданные 'ветвь') подстраховываемся.
    has_branch_col = 'Branch' in df.columns
    if not has_branch_col:
        df['Branch'] = metadata.get('ветвь', 'forward')

    # Колонка возбуждения называется X_set (новые файлы, могут быть током
    # или напряжением — единица берётся из метаданных) либо I_set_A
    # (старые файлы до появления выбора типа возбуждения — всегда ток).
    if 'X_set' in df.columns:
        excitation_col = 'X_set'
        excitation_unit = metadata.get('Единица измерения возбуждения', 'А')
        excitation_type = metadata.get('Тип возбуждения', 'current')
    else:
        excitation_col = 'I_set_A'
        excitation_unit = 'А'
        excitation_type = 'current'

    excitation_label = 'ток' if excitation_type == 'current' else 'напряжение'

    X_start = df[excitation_col].min()
    X_stop = df[excitation_col].max()

    # ---------- Расчёт погрешности ----------
    # Погрешность всегда считается относительно выходного тока датчика
    # (I_meas_A), независимо от того, чем датчик возбуждался.
    K = 1.0 / X                      # коэффициент передачи I_out / X_in
    I_sec_nom = I_nom * K            # номинальный выходной ток при I_nom

    df['I_expected_A'] = df[excitation_col] * K
    # Погрешность знаковая (п.31): по ней видно не только величину
    # расхождения, но и его направление (датчик завышает/занижает выход).
    # Все места ниже, где раньше подразумевалась неотрицательность
    # (сводный "максимум", подписи на графике), берут abs() явно, там, где
    # это действительно нужно, а не потому что оно "само так получалось".
    df['Error_percent'] = (df['I_meas_A'] - df['I_expected_A']) / I_sec_nom * 100

    # Точки не участвуют в сводной статистике погрешности по двум причинам:
    #  - Rejected — забракованы контрольными промерами (п.9, measurement.py),
    #    показание стабильно не соответствует уставке;
    #  - ManuallyExcluded — оператор вручную исключил точку из отчёта в UI
    #    (п.26); это не удаление данных — колонка обратима.
    # В обоих случаях строка остаётся в df целиком — это сырые данные.
    # Старые CSV (до Ф2/Ф3) этих колонок не имеют вовсе — тогда все точки
    # участвуют, как и раньше.
    excluded_mask = pd.Series(False, index=df.index)
    if 'Rejected' in df.columns:
        excluded_mask |= df['Rejected'].fillna(False).astype(bool)
    if 'ManuallyExcluded' in df.columns:
        excluded_mask |= df['ManuallyExcluded'].fillna(False).astype(bool)
    accepted = df[~excluded_mask]
    stats_source = accepted if not accepted.empty else df

    # ---------- Построение графиков ----------
    plt.style.use('default')
    fig, (ax1, ax2) = plt.subplots(
        2, 1, figsize=(12, 10), sharex=True,
        gridspec_kw={'height_ratios': [3, 1]},
    )
    fig.patch.set_facecolor('white')
    ax1.set_facecolor('white')
    ax2.set_facecolor('white')
    ax1.minorticks_on()
    ax2.minorticks_on()

    x_label = f"1:{int(X)}" if float(X).is_integer() else f"1:{X:.1f}"

    # Верхний график: выходной ток. Forward и reverse рисуются отдельно —
    # это две разные ветви одного и того же прохода через 0, их нельзя
    # просто сортировать вместе по excitation_col без учёта знака.
    branch_styles = {
        'forward': dict(color='steelblue', marker='o', label=f'{label} (forward) – измер.'),
        'reverse': dict(color='seagreen', marker='s', label=f'{label} (reverse) – измер.'),
    }
    for branch_name, sub in df.groupby('Branch', sort=False):
        sub = sub.sort_values(excitation_col)
        style = branch_styles.get(branch_name, dict(color='gray', marker='.', label=f'{label} ({branch_name})'))
        ax1.plot(sub[excitation_col], sub['I_meas_A'], marker=style['marker'], linestyle='-',
                  color=style['color'], markersize=4, label=style['label'])

    df_sorted_for_expected = df.sort_values(excitation_col)
    ax1.plot(df_sorted_for_expected[excitation_col], df_sorted_for_expected['I_expected_A'], '--',
              color='orange', linewidth=1.5, label=f'Ожидаемый ({x_label})')
    ax1.set_ylabel('Выходной ток датчика, А')
    ax1.set_title(f'Амплитудная характеристика датчика тока (возбуждение — {excitation_label})\n'
                  f'Диапазон {X_start}..{X_stop} {excitation_unit}')
    ax1.legend(loc='upper left')
    ax1.grid(True, which='major', linestyle='-', linewidth=0.6, alpha=0.7)
    ax1.grid(True, which='minor', linestyle=':', linewidth=0.4, alpha=0.5)

    # Нижний график: приведённая погрешность.
    # ВАЖНО: forward и reverse обе проходят через excitation_col=0, так что
    # в объединённых данных x не строго возрастает (дубли на 0, а иногда и
    # на других точках при неровном шаге). CubicSpline требует строго
    # возрастающую последовательность x, поэтому дубли усредняем перед
    # построением сплайна — сами измеренные точки (крестики) остаются
    # нетронутыми и показывают обе ветви как есть.
    x = df[excitation_col].values
    y = df['Error_percent'].values
    order = np.argsort(x)
    x_sorted, y_sorted = x[order], y[order]

    x_unique, unique_idx, counts = np.unique(x_sorted, return_index=True, return_counts=True)
    y_avg = np.array([
        y_sorted[start:start + count].mean()
        for start, count in zip(unique_idx, counts)
    ])

    if SCIPY_AVAILABLE and len(x_unique) > 3:
        cs = CubicSpline(x_unique, y_avg)
        x_smooth = np.linspace(x_unique[0], x_unique[-1], 500)
        ax2.plot(x_smooth, cs(x_smooth), '-', color='firebrick', linewidth=1.2,
                  label='Погрешность приведённая (сглаженная)')
    else:
        ax2.plot(x_unique, y_avg, '-', color='firebrick', linewidth=1.2,
                  label='Погрешность приведённая')

    ax2.plot(df[excitation_col], df['Error_percent'], 'x', color='firebrick', markersize=6, alpha=0.7)
    ax2.axhline(y=0, color='gray', linewidth=0.5)
    ax2.set_xlabel(f'Заданное возбуждение ({excitation_label}), {excitation_unit}')
    ax2.set_ylabel('Погрешность, %')
    ax2.legend(loc='upper right')
    ax2.grid(True, which='major', linestyle='-', linewidth=0.6, alpha=0.7)
    ax2.grid(True, which='minor', linestyle=':', linewidth=0.4, alpha=0.5)

    # Подписи погрешности над точками (п.30) — вертикальные (horizontal
    # текст на плотном шаге налезает сам на себя), точность 0.001 %.
    if show_error_labels:
        for xv, yv in zip(df[excitation_col].values, df['Error_percent'].values):
            ax2.annotate(f"{yv:.3f}%", xy=(xv, yv), xytext=(0, 6), textcoords='offset points',
                        rotation=90, ha='center', va='bottom', fontsize=7, color='firebrick')

    if xlim is not None:
        ax1.set_xlim(*xlim)
    if y1lim is not None:
        ax1.set_ylim(*y1lim)
    if y2lim is not None:
        ax2.set_ylim(*y2lim)

    plt.tight_layout()

    png_path: Optional[Path] = None
    if save_png:
        png_path = latest_csv.with_suffix('.png')
        plt.savefig(png_path, dpi=150, bbox_inches='tight')

    if show:
        plt.show()

    # close_fig=False используется GUI, чтобы встроить фигуру в окно
    # (FigureCanvasTkAgg); в этом случае ответственность за close — на вызове.
    if close_fig:
        plt.close(fig)

    branches_present = sorted(df['Branch'].unique().tolist())

    # max_error_percent — знаковое значение ИМЕННО в точке с наибольшим
    # расхождением ПО МОДУЛЮ (иначе "максимум" мог бы найти +0.05% там, где
    # на самом деле есть -2%, просто потому что -2 < +0.05 арифметически).
    # Знак сохраняется в результате, чтобы было видно направление отклонения.
    abs_errors = stats_source['Error_percent'].abs()
    idx_max = abs_errors.idxmax()
    max_error_percent = float(stats_source.loc[idx_max, 'Error_percent'])
    mean_error_percent = float(stats_source['Error_percent'].mean())

    stats = {
        'csv_path': latest_csv,
        'png_path': png_path,
        'label': label,
        'branches': branches_present,
        'excitation_type': excitation_type,
        'excitation_unit': excitation_unit,
        'X_start': X_start,
        'X_stop': X_stop,
        'I_nom': I_nom,
        'X': X,
        'points': len(df),
        'rejected_points': int(len(df) - len(accepted)),
        'max_error_percent': max_error_percent,
        'mean_error_percent': mean_error_percent,
        'dataframe': df,
        'figure': None if close_fig else fig,
    }
    return stats


# ----------------------------------------------------------------------
# п.22 — автопостроение графика по завершении измерительного цикла
# ----------------------------------------------------------------------

def load_and_analyze_from_params(csv_path: Path, params: dict, **kwargs) -> Optional[dict]:
    """
    Обёртка над load_and_analyze для автопостроения (п.22) — CLI и GUI
    вызывают её сразу после измерения тем же кодом, каким пользуется ручной
    режим (кнопка/подкоманда analyze, п.20), просто беря I_nom/X из уже
    известных params, а не из повторного ввода.

    Без I_nom/ratio строить не из чего (не с чем сравнивать показания) —
    это не ошибка измерения, поэтому тихо возвращаем None, а не бросаем.
    """
    I_nom = params.get('I_nom')
    X = params.get('ratio')
    if not I_nom or not X:
        return None
    return load_and_analyze(csv_path, I_nom=I_nom, X=X, **kwargs)


def metadata_i_nom_and_ratio(csv_path: Path) -> tuple:
    """
    Достаёт I_nom и коэффициент 1:X из шапки метаданных CSV (см.
    orchestrate.write_results_csv), если они там были сохранены.

    Нужно для п.20 (график из произвольного файла): при открытии старого
    CSV разумно предзаполнить поля анализа тем, с чем он снимался, а не
    оставлять то, что было в них до этого от другого файла/сессии.
    Возвращает (I_nom, X), любое из значений может быть None.
    """
    metadata = _read_metadata(Path(csv_path))
    I_nom = None
    X = None
    raw_inom = metadata.get('Номинальный первичный ток')
    if raw_inom:
        match = re.match(r'([\d.,]+)', raw_inom)
        if match:
            I_nom = float(match.group(1).replace(',', '.'))
    # orchestrate.write_results_csv пишет строку "# Коэффициент преобразования
    # 1:{X}" — единственное двоеточие в ней стоит между литеральной "1" и
    # самим X, поэтому _read_metadata (парсит по ПЕРВОМУ ':') кладёт "1" в
    # конец КЛЮЧА, а не в значение: ключ = "Коэффициент преобразования 1".
    raw_ratio = metadata.get('Коэффициент преобразования 1')
    if raw_ratio:
        match = re.match(r'([\d.,]+)', raw_ratio.strip())
        if match:
            X = float(match.group(1).replace(',', '.'))
    return I_nom, X


# ----------------------------------------------------------------------
# п.10 (BETA) — определение фактического коэффициента преобразования
# ----------------------------------------------------------------------

def estimate_ratio_from_data(df: pd.DataFrame, excitation_col: str = 'X_set') -> dict:
    """
    BETA (см. PLAN_V2.md, В-4): определяет фактический коэффициент
    преобразования 1:X по уже снятым точкам методом наименьших квадратов —
    прямая через ноль I_meas = X_set / X_actual, — и округляет его до
    ближайшего кратного 50, как того требует ТЗ.

    Рядом с округлённым значением всегда возвращается и фактическое, и
    процент расхождения между ними: для малых коэффициентов шаг 50 грубый
    (например, между 1:50 и 1:100 ничего нет), и оператор должен это видеть,
    а не получать округление молча.
    """
    excluded_mask = pd.Series(False, index=df.index)
    if 'Rejected' in df.columns:
        excluded_mask |= df['Rejected'].fillna(False).astype(bool)
    if 'ManuallyExcluded' in df.columns:
        excluded_mask |= df['ManuallyExcluded'].fillna(False).astype(bool)
    accepted = df[~excluded_mask]

    x = accepted[excitation_col].to_numpy(dtype=float)
    y = accepted['I_meas_A'].to_numpy(dtype=float)
    denom = float(np.sum(x * x))
    if denom == 0.0:
        raise ValueError("Недостаточно данных для определения коэффициента: все точки на нуле возбуждения.")
    slope = float(np.sum(x * y)) / denom
    if slope == 0.0:
        raise ValueError("Не удалось определить коэффициент: измеренный ток нулевой во всех точках.")

    X_actual = abs(1.0 / slope)
    X_rounded = max(50.0, round(X_actual / 50.0) * 50.0)
    discrepancy_percent = abs(X_rounded - X_actual) / X_actual * 100.0

    return {
        'X_actual': X_actual,
        'X_rounded': X_rounded,
        'discrepancy_percent': discrepancy_percent,
    }


# ----------------------------------------------------------------------
# п.21 — экспорт в XLSX
# ----------------------------------------------------------------------

def export_xlsx(csv_path: Path, xlsx_path: Optional[Path] = None) -> Path:
    """
    Экспорт результатов измерения в XLSX (п.21): лист "Данные" — полная
    копия таблицы CSV с закреплённой шапкой и подобранными по содержимому
    ширинами колонок, лист "Метаданные" — те же '#'-строки шапки CSV, но в
    виде пар параметр/значение.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path) if xlsx_path is not None else csv_path.with_suffix('.xlsx')

    metadata = _read_metadata(csv_path)
    df = pd.read_csv(csv_path, comment='#')

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Данные"
    ws_data.append(list(df.columns))
    for cell in ws_data[1]:
        cell.font = Font(bold=True)
    for row in df.itertuples(index=False):
        ws_data.append(list(row))
    ws_data.freeze_panes = "A2"
    for i, col in enumerate(df.columns, start=1):
        content_width = df[col].astype(str).map(len).max() if len(df) else 0
        width = max(len(str(col)), content_width) + 2
        ws_data.column_dimensions[get_column_letter(i)].width = min(width, 40)

    ws_meta = wb.create_sheet("Метаданные")
    ws_meta.append(["Параметр", "Значение"])
    for cell in ws_meta[1]:
        cell.font = Font(bold=True)
    for key, value in metadata.items():
        ws_meta.append([key, value])
    ws_meta.column_dimensions['A'].width = 42
    ws_meta.column_dimensions['B'].width = 60
    ws_meta.freeze_panes = "A2"

    wb.save(xlsx_path)
    return xlsx_path


# ----------------------------------------------------------------------
# п.26 — правка точек: сохранение исключений, инверсия входа (пост-обработка)
# ----------------------------------------------------------------------

def save_dataframe_with_metadata(csv_path: Path, df: pd.DataFrame) -> None:
    """
    Перезаписывает CSV тем же '#'-заголовком метаданных, что был в файле, с
    изменённой таблицей данных — например, после того как оператор отметил
    точки как ManuallyExcluded в UI (п.26). Само исключение обратимо: строки
    не удаляются, только помечаются, поэтому "сохранить" здесь — это просто
    "переписать колонку", а не потеря сырых данных.
    """
    csv_path = Path(csv_path)
    header_lines = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        for line in f:
            if not line.startswith('#'):
                break
            header_lines.append(line)

    with open(csv_path, 'w', encoding='utf-8') as f:
        f.writelines(header_lines)
        df.to_csv(f, index=False)


def apply_invert_input(csv_path: Path, output_path: Optional[Path] = None) -> Path:
    """
    Пост-обработка, перенесённая из измерительного цикла (см. PLAN_V2.md,
    §0.1/Ф0): раньше invert_input подделывал знак X_set прямо во время
    записи данных. Теперь это явная операция над уже снятым файлом, и она
    пишет результат в НОВЫЙ CSV (с пометкой в шапке), а не поверх исходного —
    сырые данные измерения не должны зависеть от того, догадался ли кто-то
    потом инвертировать их обратно.

    НЕ РЕКОМЕНДУЕТСЯ к использованию (см. документацию) — годится только
    когда датчик физически подключён в обратной полярности, а
    перекоммутировать его на стенде нельзя.
    """
    csv_path = Path(csv_path)
    output_path = Path(output_path) if output_path is not None else csv_path.with_name(
        csv_path.stem + "_inverted.csv")

    header_lines = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        for line in f:
            if not line.startswith('#'):
                break
            header_lines.append(line)

    df = pd.read_csv(csv_path, comment='#')
    for col in ('X_set', 'X_real'):
        if col in df.columns:
            df[col] = -df[col]

    with open(output_path, 'w', encoding='utf-8') as f:
        f.writelines(header_lines)
        f.write("# Пост-обработка: знак заданного возбуждения инвертирован (invert_input, "
                "НЕ РЕКОМЕНДУЕТСЯ, см. документацию)\n")
        f.write("#\n")
        df.to_csv(f, index=False)

    return output_path
