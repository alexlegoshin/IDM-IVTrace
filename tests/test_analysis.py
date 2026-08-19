import time

import pandas as pd
import pytest

from analysis import (
    load_and_analyze, find_latest_csv, _read_metadata,
    load_and_analyze_from_params, metadata_i_nom_and_ratio, metadata_zero_offset,
    estimate_ratio_from_data, export_xlsx,
    save_dataframe_with_metadata, apply_invert_input,
)


def _write_csv(path, extra_header_lines=(), rows=None, excitation_col='X_set'):
    rows = rows if rows is not None else [
        ('forward', 0, 0.0),
        ('forward', 5, 0.005),
        ('reverse', 0, 0.0),
        ('reverse', -5, -0.0051),
    ]
    with open(path, 'w', encoding='utf-8') as f:
        f.write("# Датчик: TestSensor\n")
        for line in extra_header_lines:
            f.write(line + "\n")
        f.write("#\n")
        f.write(f"Timestamp,Branch,{excitation_col},I_meas_A\n")
        for branch, x, i in rows:
            f.write(f"2026-01-01T00:00:00,{branch},{x},{i}\n")


def _write_csv_y_meas(path, extra_header_lines=(), rows=None):
    """Как _write_csv, но новой (ось А-1) колонкой Y_meas вместо I_meas_A."""
    rows = rows if rows is not None else [
        ('forward', 0, 0.0),
        ('forward', 5, 0.005),
        ('reverse', 0, 0.0),
        ('reverse', -5, -0.0051),
    ]
    with open(path, 'w', encoding='utf-8') as f:
        f.write("# Датчик: TestSensor\n")
        for line in extra_header_lines:
            f.write(line + "\n")
        f.write("#\n")
        f.write("Timestamp,Branch,X_set,Y_meas\n")
        for branch, x, y in rows:
            f.write(f"2026-01-01T00:00:00,{branch},{x},{y}\n")


# ----------------------------------------------------------------------
# find_latest_csv
# ----------------------------------------------------------------------

def test_find_latest_csv_returns_most_recently_modified(tmp_path):
    old = tmp_path / "IVtrace_a_20260101_000000.csv"
    new = tmp_path / "IVtrace_b_20260102_000000.csv"
    _write_csv(old)
    time.sleep(0.05)
    _write_csv(new)

    assert find_latest_csv(tmp_path) == new


def test_find_latest_csv_missing_dir_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        find_latest_csv(tmp_path / "nope")


def test_find_latest_csv_no_matching_files_raises(tmp_path):
    (tmp_path / "unrelated.csv").write_text("x", encoding='utf-8')
    with pytest.raises(FileNotFoundError):
        find_latest_csv(tmp_path)


# ----------------------------------------------------------------------
# _read_metadata
# ----------------------------------------------------------------------

def test_read_metadata_parses_hash_lines(tmp_path):
    csv_path = tmp_path / "IVtrace_x_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
    ])
    meta = _read_metadata(csv_path)
    assert meta['Датчик'] == 'TestSensor'
    assert meta['Тип возбуждения'] == 'current'
    assert meta['Единица измерения возбуждения'] == 'A'


# ----------------------------------------------------------------------
# load_and_analyze
# ----------------------------------------------------------------------

def test_graph_excludes_rejected_but_keeps_in_dataframe_and_axis_without_gost(tmp_path):
    """
    Ш5: бракованная точка НЕ строится на графике (баг п.9), но остаётся в
    df/return (сырые данные для протокола); ось Y нижнего графика — просто
    «Погрешность, %» без «ГОСТ»/«приведённая» (баг п.19); ожидаемая прямая —
    на фоне (zorder ниже измеренных, баг п.5).
    """
    import matplotlib.pyplot as plt

    csv_path = tmp_path / "IVtrace_g_20260101_000000.csv"
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("# Датчик: G\n#\n")
        f.write("Timestamp,Branch,X_set,X_real,Y_meas,Y_unit,Rejected,RejectReason,PolarityMismatch\n")
        f.write("t,zero,0,0,0.0,A,False,,False\n")
        f.write("t,forward,10,10,0.10,A,False,,False\n")
        f.write('t,forward,20,20,0.55,A,True,"brak",False\n')  # бракованная
        f.write("t,forward,30,30,0.30,A,False,,False\n")

    stats = load_and_analyze(csv_path, I_nom=100.0, X=100.0, save_png=False, close_fig=False)
    fig = stats["figure"]
    ax1, ax2 = fig.axes[0], fig.axes[1]

    # ось без ГОСТ
    assert ax2.get_ylabel() == "Погрешность, %"
    assert "ГОСТ" not in ax2.get_ylabel()

    # брак остаётся в данных/статистике
    assert stats["rejected_points"] == 1
    assert len(stats["dataframe"]) == 4

    # но НЕ на графике: бракованная X=20 не входит ни в одну измеренную линию
    plotted_x, expected_zorders, measured_zorders = set(), [], []
    for line in ax1.get_lines():
        if line.get_linestyle() == "--":  # ожидаемая прямая
            expected_zorders.append(line.get_zorder())
            continue
        measured_zorders.append(line.get_zorder())
        plotted_x.update(line.get_xdata().tolist())
    assert 20.0 not in plotted_x
    assert 30.0 in plotted_x and 10.0 in plotted_x
    # ожидаемая — на фоне (ниже измеренных)
    assert max(expected_zorders) < min(measured_zorders)

    plt.close(fig)


def test_load_and_analyze_computes_expected_error(tmp_path):
    csv_path = tmp_path / "IVtrace_x_20260101_000000.csv"
    # K = I_meas / X_set идеально линеен -> ожидаем нулевую (или почти нулевую) погрешность.
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
    ], rows=[
        ('forward', 0, 0.0),
        ('forward', 10, 0.1),   # K=0.01
        ('reverse', 0, 0.0),
        ('reverse', -10, -0.1),
    ])

    # I_nom=1000 А, X=100 (1:100) -> I_sec_nom = 10 А -> K расчётный = I_sec_nom/I_nom = 0.01,
    # что совпадает с реальным K измерений -> погрешность около нуля.
    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=True, show=False)

    assert stats['label'] == 'TestSensor'
    assert stats['branches'] == ['forward', 'reverse']
    assert stats['points'] == 4
    assert stats['max_error_percent'] == pytest.approx(0.0, abs=1e-9)
    assert stats['png_path'].exists()


def test_load_and_analyze_detects_nonzero_error(tmp_path):
    csv_path = tmp_path / "IVtrace_x_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
    ], rows=[
        ('forward', 0, 0.0),
        ('forward', 10, 0.11),  # 10% выше ожидаемого
    ])

    # I_nom=1000, X=100 -> I_sec_nom=10; expected=0.1; измерено=0.11 -> |0.01|/10*100 = 0.1%
    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    assert stats['max_error_percent'] == pytest.approx(0.1, abs=1e-9)


def test_load_and_analyze_uses_x_real_not_x_set_when_turns_applied(tmp_path):
    # Баг-репорт: раньше Y_expected/Error_percent считались от голой
    # уставки X_set, игнорируя число витков. turns=10 -> при X_set=1
    # реальный вход датчика X_real=10. Датчик 1:100, I_nom=1000 ->
    # Y_sec_nom=10, ожидаемый выход при X_real=10 равен 0.1 А — ровно то,
    # что измерено, погрешность должна быть нулевой. Со старым багом
    # (счёт от X_set=1) график решил бы, что ожидание — 0.01 А, и приписал
    # бы исправному датчику несуществующую ~1%-ную погрешность.
    csv_path = tmp_path / "IVtrace_turns_20260101_000000.csv"
    with open(csv_path, 'w', encoding='utf-8') as f:
        f.write("# Датчик: TestSensor\n")
        f.write("# Тип возбуждения: current\n")
        f.write("# Единица измерения возбуждения: A\n")
        f.write("# Число витков через датчик: 10 (реальный вход = X_set × витки, см. колонку X_real)\n")
        f.write("#\n")
        f.write("Timestamp,Branch,X_set,X_real,I_meas_A\n")
        f.write("2026-01-01T00:00:00,forward,0,0,0.0\n")
        f.write("2026-01-01T00:00:00,forward,1,10,0.1\n")

    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    assert stats['max_error_percent'] == pytest.approx(0.0, abs=1e-9)


def test_load_and_analyze_falls_back_to_x_set_without_x_real_column(tmp_path):
    # Старые CSV без X_real (до появления turns) — поведение как раньше.
    csv_path = tmp_path / "IVtrace_noturns_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
    ], rows=[
        ('forward', 0, 0.0),
        ('forward', 10, 0.1),
    ])
    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    assert stats['max_error_percent'] == pytest.approx(0.0, abs=1e-9)


# ----------------------------------------------------------------------
# zero_offset (feature "offset нуля") — читается из шапки CSV либо
# передаётся явным параметром (переопределяет шапку)
# ----------------------------------------------------------------------

def test_load_and_analyze_applies_zero_offset_from_csv_metadata(tmp_path):
    csv_path = tmp_path / "IVtrace_offset_20260101_000000.csv"
    # Смещение +0.01 сидит в КАЖДОМ показании, включая X=0 (физически так и
    # проявляется реальный сдвиг нуля датчика — см. X=0 ниже: 0.01, а не 0):
    # истинное значение при X=10 — 0.1, но сырое показание 0.11.
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
        "# Смещение нуля: 0.01 А (вычитается из Y_meas при расчёте погрешности — см. analysis.py)",
    ], rows=[
        ('forward', 0, 0.01),
        ('forward', 10, 0.11),
    ])

    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    assert stats['max_error_percent'] == pytest.approx(0.0, abs=1e-9)


def test_load_and_analyze_explicit_zero_offset_overrides_metadata(tmp_path):
    csv_path = tmp_path / "IVtrace_offset2_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
        "# Смещение нуля: 0.01 А (вычитается из Y_meas при расчёте погрешности — см. analysis.py)",
    ], rows=[
        ('forward', 0, 0.01),
        ('forward', 10, 0.11),
    ])

    # Явный zero_offset=0.0 должен ПЕРЕОПРЕДЕЛИТЬ метаданные файла (0.01) —
    # возвращается прежняя (без поправки) ошибка: при X=0 (0.01-0)/10*100=0.1%,
    # при X=10 (0.11-0.1)/10*100=0.1% — оба одинаковы по модулю.
    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False,
                             zero_offset=0.0)
    assert stats['max_error_percent'] == pytest.approx(0.1, abs=1e-9)


def test_load_and_analyze_without_zero_offset_metadata_behaves_as_before(tmp_path):
    csv_path = tmp_path / "IVtrace_nooffset_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
    ], rows=[
        ('forward', 0, 0.0),
        ('forward', 10, 0.11),
    ])
    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    assert stats['max_error_percent'] == pytest.approx(0.1, abs=1e-9)


def test_metadata_zero_offset_parses_leading_number_ignoring_unit_and_note(tmp_path):
    csv_path = tmp_path / "IVtrace_meta_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Смещение нуля: 0.5 А (вычитается из Y_meas при расчёте погрешности — см. analysis.py)",
    ])
    assert metadata_zero_offset(csv_path) == pytest.approx(0.5)


def test_metadata_zero_offset_handles_negative_value(tmp_path):
    csv_path = tmp_path / "IVtrace_meta_neg_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Смещение нуля: -0.25 А (вычитается из Y_meas при расчёте погрешности — см. analysis.py)",
    ])
    assert metadata_zero_offset(csv_path) == pytest.approx(-0.25)


def test_metadata_zero_offset_missing_returns_none(tmp_path):
    csv_path = tmp_path / "IVtrace_meta_none_20260101_000000.csv"
    _write_csv(csv_path)
    assert metadata_zero_offset(csv_path) is None


def test_estimate_ratio_from_data_applies_zero_offset(tmp_path):
    # Y_meas = X_set/1000 + смещение 0.5 -> без поправки наклон искажён.
    df = pd.DataFrame([
        {'X_set': 100.0, 'Y_meas': 100.0 / 1000 + 0.5},
        {'X_set': 200.0, 'Y_meas': 200.0 / 1000 + 0.5},
        {'X_set': 300.0, 'Y_meas': 300.0 / 1000 + 0.5},
    ])
    result = estimate_ratio_from_data(df, zero_offset=0.5)
    assert result['X_actual'] == pytest.approx(1000.0, rel=1e-6)


def test_load_and_analyze_excludes_rejected_points_from_stats_but_keeps_them_in_dataframe(tmp_path):
    # Забракованные контрольными промерами точки (п.9, measurement.py)
    # остаются в сырых данных, но не должны портить сводную погрешность.
    csv_path = tmp_path / "IVtrace_rejected_20260101_000000.csv"
    with open(csv_path, 'w', encoding='utf-8') as f:
        f.write("# Датчик: TestSensor\n")
        f.write("# Тип возбуждения: current\n")
        f.write("# Единица измерения возбуждения: A\n")
        f.write("#\n")
        f.write("Timestamp,Branch,X_set,I_meas_A,Rejected\n")
        f.write("2026-01-01T00:00:00,forward,0,0.0,False\n")
        f.write("2026-01-01T00:00:00,forward,10,0.1,False\n")   # K=0.01, точно ожидаемое
        f.write("2026-01-01T00:00:00,forward,20,999.0,True\n")  # заведомо бракованное показание

    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)

    assert stats['points'] == 3          # сырые данные — все три строки
    assert stats['rejected_points'] == 1
    # Без исключения брака max_error_percent улетел бы в тысячи процентов.
    assert stats['max_error_percent'] == pytest.approx(0.0, abs=1e-9)
    assert len(stats['dataframe']) == 3  # точка осталась в df целиком, не удалена


def test_load_and_analyze_without_rejected_column_uses_all_points(tmp_path):
    # Обратная совместимость: CSV до Ф2 без колонки Rejected — все точки участвуют.
    csv_path = tmp_path / "IVtrace_no_rejected_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
    ], rows=[('forward', 0, 0.0), ('forward', 10, 0.1)])

    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    assert stats['rejected_points'] == 0
    assert stats['points'] == 2


def test_load_and_analyze_empty_csv_raises(tmp_path):
    csv_path = tmp_path / "IVtrace_x_20260101_000000.csv"
    _write_csv(csv_path, rows=[])
    with pytest.raises(ValueError):
        load_and_analyze(csv_path, I_nom=100.0, X=10.0)


def test_load_and_analyze_backward_compat_old_i_set_a_column(tmp_path):
    """Старые CSV без X_set (до появления выбора типа возбуждения) должны читаться как ток."""
    csv_path = tmp_path / "IVtrace_old_20260101_000000.csv"
    _write_csv(csv_path, rows=[
        ('forward', 0, 0.0),
        ('forward', 5, 0.05),
    ], excitation_col='I_set_A')

    stats = load_and_analyze(csv_path, I_nom=100.0, X=10.0, save_png=False, show=False)
    assert stats['excitation_type'] == 'current'
    assert stats['excitation_unit'] == 'А'


def test_load_and_analyze_missing_branch_column_defaults_to_forward(tmp_path):
    csv_path = tmp_path / "IVtrace_nobranch_20260101_000000.csv"
    with open(csv_path, 'w', encoding='utf-8') as f:
        f.write("# Датчик: OldSensor\n")
        f.write("#\n")
        f.write("Timestamp,X_set,I_meas_A\n")
        f.write("2026-01-01T00:00:00,0,0.0\n")
        f.write("2026-01-01T00:00:00,5,0.05\n")

    stats = load_and_analyze(csv_path, I_nom=100.0, X=10.0, save_png=False, show=False)
    assert stats['branches'] == ['forward']


# ----------------------------------------------------------------------
# п.31 — знаковая погрешность
# ----------------------------------------------------------------------

def test_error_percent_keeps_sign_of_deviation(tmp_path):
    csv_path = tmp_path / "IVtrace_signed_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
    ], rows=[
        ('forward', 0, 0.0),
        ('forward', 10, 0.09),   # ниже ожидаемого (0.1) -> отрицательная погрешность
    ])
    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    # measured < expected -> знак минус, не abs().
    assert stats['max_error_percent'] == pytest.approx(-0.1, abs=1e-9)
    assert (stats['dataframe']['Error_percent'] < 0).any()


def test_max_error_percent_picks_largest_magnitude_keeping_its_sign(tmp_path):
    csv_path = tmp_path / "IVtrace_signmix_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
    ], rows=[
        ('forward', 0, 0.0),
        ('forward', 10, 0.1005),   # expected 0.1 -> +0.05% - маленькое положительное отклонение
        ('forward', 20, 0.1),      # expected 0.2 -> (0.1-0.2)/10*100 = -1.0% - крупное отрицательное
    ])
    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    assert stats['max_error_percent'] < 0
    assert abs(stats['max_error_percent']) > 0.5


# ----------------------------------------------------------------------
# п.26 — ManuallyExcluded исключается из статистики так же, как Rejected
# ----------------------------------------------------------------------

def test_manually_excluded_points_removed_from_stats_but_kept_in_dataframe(tmp_path):
    csv_path = tmp_path / "IVtrace_excl_20260101_000000.csv"
    with open(csv_path, 'w', encoding='utf-8') as f:
        f.write("# Датчик: TestSensor\n")
        f.write("# Тип возбуждения: current\n")
        f.write("# Единица измерения возбуждения: A\n")
        f.write("#\n")
        f.write("Timestamp,Branch,X_set,I_meas_A,ManuallyExcluded\n")
        f.write("2026-01-01T00:00:00,forward,0,0.0,False\n")
        f.write("2026-01-01T00:00:00,forward,10,0.1,False\n")
        f.write("2026-01-01T00:00:00,forward,20,999.0,True\n")

    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    assert stats['points'] == 3
    assert stats['rejected_points'] == 1
    assert stats['max_error_percent'] == pytest.approx(0.0, abs=1e-9)
    assert len(stats['dataframe']) == 3


def test_manually_excluded_points_not_drawn_on_graph(tmp_path):
    """
    Баг-репорт: после «Исключить + Сохранить и перестроить» точка всё равно
    строилась. ManuallyExcluded должна исключаться не только из статистики, но
    и из ПОСТРОЕНИЯ (оставаясь в df/CSV).
    """
    csv_path = tmp_path / "IVtrace_mex_20260101_000000.csv"
    with open(csv_path, 'w', encoding='utf-8') as f:
        f.write("# Датчик: T\n#\n")
        f.write("Timestamp,Branch,X_set,X_real,Y_meas,Y_unit,Rejected,RejectReason,PolarityMismatch,ManuallyExcluded\n")
        f.write("t,zero,0,0,0.0,A,False,,False,False\n")
        f.write("t,forward,10,10,0.10,A,False,,False,False\n")
        f.write("t,forward,20,20,0.30,A,False,,False,True\n")   # исключена вручную
        f.write("t,forward,30,30,0.30,A,False,,False,False\n")

    # close_fig=False -> объектная Figure (без pyplot), её отдельно закрывать
    # не нужно (соберётся сборщиком мусора).
    stats = load_and_analyze(csv_path, I_nom=100.0, X=100.0, save_png=False, close_fig=False)
    plotted = set()
    for line in stats['figure'].axes[0].get_lines():
        if line.get_linestyle() == '--':
            continue
        plotted.update(line.get_xdata().tolist())
    assert 20.0 not in plotted          # исключённая не рисуется
    assert 30.0 in plotted              # обычная рисуется
    assert len(stats['dataframe']) == 4  # но остаётся в данных


# ----------------------------------------------------------------------
# п.30 — подписи погрешности (смоук: не падает, PNG создаётся)
# ----------------------------------------------------------------------

def test_show_error_labels_does_not_crash_and_saves_png(tmp_path):
    csv_path = tmp_path / "IVtrace_labels_20260101_000000.csv"
    _write_csv(csv_path)
    stats = load_and_analyze(csv_path, I_nom=100.0, X=10.0, save_png=True, show=False,
                             show_error_labels=True)
    assert stats['png_path'].exists()


# ----------------------------------------------------------------------
# п.36 — диапазоны осей (смоук)
# ----------------------------------------------------------------------

def test_axis_limits_do_not_crash_and_save_png(tmp_path):
    csv_path = tmp_path / "IVtrace_lims_20260101_000000.csv"
    _write_csv(csv_path)
    stats = load_and_analyze(csv_path, I_nom=100.0, X=10.0, save_png=True, show=False,
                             xlim=(-6, 6), y1lim=(-0.01, 0.01), y2lim=(-5, 5))
    assert stats['png_path'].exists()


# ----------------------------------------------------------------------
# load_and_analyze_from_params (п.22 — автопостроение по завершении цикла)
# ----------------------------------------------------------------------

def test_load_and_analyze_from_params_returns_none_without_inom_or_ratio(tmp_path):
    csv_path = tmp_path / "IVtrace_noparam_20260101_000000.csv"
    _write_csv(csv_path)
    assert load_and_analyze_from_params(csv_path, {}, save_png=False, show=False) is None
    assert load_and_analyze_from_params(csv_path, {'I_nom': 100.0}, save_png=False, show=False) is None
    assert load_and_analyze_from_params(csv_path, {'ratio': 10.0}, save_png=False, show=False) is None


def test_load_and_analyze_from_params_builds_stats_when_both_present(tmp_path):
    csv_path = tmp_path / "IVtrace_bothparam_20260101_000000.csv"
    _write_csv(csv_path)
    stats = load_and_analyze_from_params(csv_path, {'I_nom': 100.0, 'ratio': 10.0},
                                         save_png=False, show=False)
    assert stats is not None
    assert stats['I_nom'] == 100.0
    assert stats['X'] == 10.0


# ----------------------------------------------------------------------
# metadata_i_nom_and_ratio (п.20 — предзаполнение при открытии произвольного файла)
# ----------------------------------------------------------------------

def test_metadata_i_nom_and_ratio_parses_written_header(tmp_path):
    csv_path = tmp_path / "IVtrace_meta_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Номинальный первичный ток: 150.0 А",
        "# Коэффициент преобразования 1:1500.0",
    ])
    I_nom, X, excitation_type = metadata_i_nom_and_ratio(csv_path)
    assert I_nom == 150.0
    assert X == 1500.0
    assert excitation_type == 'current'


def test_metadata_i_nom_and_ratio_missing_returns_none_none(tmp_path):
    csv_path = tmp_path / "IVtrace_nometa_20260101_000000.csv"
    _write_csv(csv_path)
    I_nom, X, excitation_type = metadata_i_nom_and_ratio(csv_path)
    assert I_nom is None
    assert X is None
    assert excitation_type == 'current'


def test_metadata_i_nom_and_ratio_reads_voltage_label(tmp_path):
    # Баг-репорт: при возбуждении напряжением шапка пишет "Номинальное
    # первичное напряжение", а не "...первичный ток" — обе подписи должны
    # распознаваться.
    csv_path = tmp_path / "IVtrace_v_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=[
        "# Тип возбуждения: voltage",
        "# Номинальное первичное напряжение: 50.0 В",
        "# Коэффициент преобразования 1:100.0",
    ])
    I_nom, X, excitation_type = metadata_i_nom_and_ratio(csv_path)
    assert I_nom == 50.0
    assert X == 100.0
    assert excitation_type == 'voltage'


# ----------------------------------------------------------------------
# estimate_ratio_from_data (п.10, BETA)
# ----------------------------------------------------------------------

def test_estimate_ratio_from_data_uses_x_real_when_present():
    # Баг-репорт: раньше коэффициент всегда оценивался по голой уставке
    # X_set, игнорируя витки. X_real = X_set * 10 (turns=10); истинный
    # коэффициент 1:1500 определён относительно РЕАЛЬНОГО входа
    # (I_meas = X_real / 1500). Со старым багом (счёт от X_set) результат
    # получился бы в 10 раз меньше настоящего (150 вместо 1500).
    df = pd.DataFrame({
        'X_set': [0.0, 15.0, 30.0, 75.0, 150.0],
        'X_real': [0.0, 150.0, 300.0, 750.0, 1500.0],
        'I_meas_A': [0.0, 0.1, 0.2, 0.5, 1.0],
    })
    result = estimate_ratio_from_data(df)
    assert result['X_actual'] == pytest.approx(1500.0, rel=1e-6)


def test_estimate_ratio_from_data_recovers_known_ratio():
    # I_meas = X_set / 1500, точно без шума -> МНК должен вернуть ровно 1500.
    df = pd.DataFrame({
        'X_set': [0.0, 150.0, 300.0, 750.0, 1500.0],
        'I_meas_A': [0.0, 0.1, 0.2, 0.5, 1.0],
    })
    result = estimate_ratio_from_data(df)
    assert result['X_actual'] == pytest.approx(1500.0, rel=1e-6)
    assert result['X_rounded'] == pytest.approx(1500.0)
    assert result['discrepancy_percent'] == pytest.approx(0.0, abs=1e-6)


def test_estimate_ratio_from_data_rounds_to_nearest_multiple_of_25():
    # I_meas = X_set / 1523 -> фактический коэффициент не кратен 25.
    df = pd.DataFrame({
        'X_set': [0.0, 1000.0],
        'I_meas_A': [0.0, 1000.0 / 1523.0],
    })
    result = estimate_ratio_from_data(df)
    assert result['X_actual'] == pytest.approx(1523.0, rel=1e-3)
    assert result['X_rounded'] == 1525.0
    assert result['discrepancy_percent'] > 0


def test_estimate_ratio_from_data_excludes_rejected_and_manually_excluded_rows():
    df = pd.DataFrame({
        'X_set': [0.0, 150.0, 300.0],
        'I_meas_A': [0.0, 0.1, 999.0],       # третья строка - явный выброс
        'Rejected': [False, False, True],
    })
    result = estimate_ratio_from_data(df)
    assert result['X_actual'] == pytest.approx(1500.0, rel=1e-6)


def test_estimate_ratio_from_data_raises_when_all_points_at_zero():
    df = pd.DataFrame({'X_set': [0.0, 0.0], 'I_meas_A': [0.0, 0.0]})
    with pytest.raises(ValueError):
        estimate_ratio_from_data(df)


# ----------------------------------------------------------------------
# export_xlsx (п.21)
# ----------------------------------------------------------------------

def test_export_xlsx_creates_file_with_data_and_metadata_sheets(tmp_path):
    from openpyxl import load_workbook

    csv_path = tmp_path / "IVtrace_xlsx_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=["# Тип возбуждения: current"])

    xlsx_path = export_xlsx(csv_path)
    assert xlsx_path == csv_path.with_suffix('.xlsx')
    assert xlsx_path.exists()

    wb = load_workbook(xlsx_path)
    assert "Данные" in wb.sheetnames
    assert "Метаданные" in wb.sheetnames
    ws = wb["Данные"]
    assert ws['A1'].value == 'Timestamp'
    assert ws.freeze_panes == "A2"

    ws_meta = wb["Метаданные"]
    meta_values = [row[0].value for row in ws_meta.iter_rows(min_row=2)]
    assert 'Датчик' in meta_values


def test_export_xlsx_accepts_explicit_output_path(tmp_path):
    csv_path = tmp_path / "IVtrace_xlsx2_20260101_000000.csv"
    _write_csv(csv_path)
    custom_path = tmp_path / "custom_name.xlsx"

    result = export_xlsx(csv_path, xlsx_path=custom_path)
    assert result == custom_path
    assert custom_path.exists()


def test_export_xlsx_survives_rejected_rows_with_nan(tmp_path):
    """
    Регрессия (баг-репорт «XLSX не сохраняется»): реальные измерения содержат
    бракованные/сбойные точки — NaN в Y_meas, пустой RejectReason, bool-колонки.
    Раньше подбор ширин `astype(str).map(len)` падал на NaN под pandas 3.x и
    ронял весь экспорт. Теперь файл должен сохраниться, NaN -> пустая ячейка.
    """
    from openpyxl import load_workbook

    csv_path = tmp_path / "IVtrace_rej_20260101_000000.csv"
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("# Датчик: TestSensor\n#\n")
        f.write("Timestamp,Branch,X_set,X_real,Y_meas,Y_unit,Rejected,RejectReason,PolarityMismatch\n")
        f.write("2026-01-01T00:00:00,zero,0.0,0.0,0.001,A,False,,False\n")
        f.write("2026-01-01T00:00:01,forward,10.0,10.0,0.099,A,False,,False\n")
        # бракованная точка: Y_meas пустой (NaN), RejectReason заполнен
        f.write('2026-01-01T00:00:02,forward,20.0,20.0,,A,True,"погрешность 20% > 1%",False\n')

    xlsx_path = export_xlsx(csv_path)
    assert xlsx_path.exists()

    wb = load_workbook(xlsx_path)
    ws = wb["Данные"]
    # строка с NaN Y_meas записана, пропуск -> пустая ячейка (None), не строка "nan"
    y_meas_col = [c.value for c in ws["E"]]
    assert y_meas_col[0] == "Y_meas"
    assert y_meas_col[3] is None  # 3-я точка данных (row 4): Y_meas был пустым
    # bool-колонка Rejected сохранена как значение, а не упала
    rejected_col = [c.value for c in ws["G"]]
    assert rejected_col[3] in (True, "True", 1)


# ----------------------------------------------------------------------
# save_dataframe_with_metadata (п.26 — сохранение исключений точек)
# ----------------------------------------------------------------------

def test_save_dataframe_with_metadata_preserves_header_and_updates_data(tmp_path):
    csv_path = tmp_path / "IVtrace_save_20260101_000000.csv"
    _write_csv(csv_path, extra_header_lines=["# Тип возбуждения: current"])

    df = pd.read_csv(csv_path, comment='#')
    df['ManuallyExcluded'] = [False] * len(df)
    df.loc[0, 'ManuallyExcluded'] = True
    save_dataframe_with_metadata(csv_path, df)

    metadata = _read_metadata(csv_path)
    assert metadata['Датчик'] == 'TestSensor'
    assert metadata['Тип возбуждения'] == 'current'

    reloaded = pd.read_csv(csv_path, comment='#')
    assert reloaded.loc[0, 'ManuallyExcluded'] == True
    assert len(reloaded) == len(df)


# ----------------------------------------------------------------------
# apply_invert_input (п.26 — пост-обработка, перенесено из Ф0)
# ----------------------------------------------------------------------

def test_apply_invert_input_flips_x_set_sign_in_new_file(tmp_path):
    csv_path = tmp_path / "IVtrace_inv_20260101_000000.csv"
    _write_csv(csv_path, rows=[('forward', 0, 0.0), ('forward', 5, 0.05), ('reverse', -5, -0.05)])

    output_path = apply_invert_input(csv_path)

    assert output_path == csv_path.with_name(csv_path.stem + "_inverted.csv")
    assert output_path.exists()
    assert csv_path.exists()  # исходный файл не тронут

    original = pd.read_csv(csv_path, comment='#')
    inverted = pd.read_csv(output_path, comment='#')
    assert list(inverted['X_set']) == [-v for v in original['X_set']]

    metadata = _read_metadata(output_path)
    assert 'invert_input' in metadata.get('Пост-обработка', '')


def test_apply_invert_input_accepts_explicit_output_path(tmp_path):
    csv_path = tmp_path / "IVtrace_inv2_20260101_000000.csv"
    _write_csv(csv_path)
    custom_path = tmp_path / "custom_inverted.csv"

    result = apply_invert_input(csv_path, output_path=custom_path)
    assert result == custom_path
    assert custom_path.exists()


# ----------------------------------------------------------------------
# Ось А-1 (PLAN_V2.md) — выход датчика: колонка Y_meas + метаданные
# output_type/output_unit, независимо от типа возбуждения
# ----------------------------------------------------------------------

def test_load_and_analyze_reads_y_meas_column_when_present(tmp_path):
    csv_path = tmp_path / "IVtrace_ymeas_20260101_000000.csv"
    _write_csv_y_meas(csv_path, extra_header_lines=[
        "# Тип возбуждения: current",
        "# Единица измерения возбуждения: A",
        "# Тип выхода датчика: voltage",
        "# Единица измерения выхода: V",
    ], rows=[('forward', 0, 0.0), ('forward', 10, 0.1)])

    stats = load_and_analyze(csv_path, I_nom=1000.0, X=100.0, save_png=False, show=False)
    assert stats['output_type'] == 'voltage'
    assert stats['output_unit'] == 'V'
    assert stats['max_error_percent'] == pytest.approx(0.0, abs=1e-9)
    assert 'Y_meas' in stats['dataframe'].columns
    assert 'Y_expected' in stats['dataframe'].columns


def test_load_and_analyze_without_y_meas_defaults_output_type_to_current(tmp_path):
    # Старые CSV (до этого пункта плана) не знают про output_type вовсе.
    csv_path = tmp_path / "IVtrace_legacy_out_20260101_000000.csv"
    _write_csv(csv_path)
    stats = load_and_analyze(csv_path, I_nom=100.0, X=10.0, save_png=False, show=False)
    assert stats['output_type'] == 'current'
    assert stats['output_unit'] == 'А'


def test_estimate_ratio_from_data_reads_y_meas_column_when_present():
    df = pd.DataFrame({
        'X_set': [0.0, 150.0, 1500.0],
        'Y_meas': [0.0, 0.1, 1.0],
    })
    result = estimate_ratio_from_data(df)
    assert result['X_actual'] == pytest.approx(1500.0, rel=1e-6)


def test_estimate_ratio_from_data_still_reads_legacy_i_meas_a_column():
    df = pd.DataFrame({
        'X_set': [0.0, 150.0, 1500.0],
        'I_meas_A': [0.0, 0.1, 1.0],
    })
    result = estimate_ratio_from_data(df)
    assert result['X_actual'] == pytest.approx(1500.0, rel=1e-6)
